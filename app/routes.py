from flask import Flask, render_template, request, send_file
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/download_excel", methods=["POST"])
def download_excel():
    # Retrieve data from the request
    data = request.get_json()
    serial_number = data.get("serialNumber")
    phantom_count = int(data.get("phantomCount", 1))

    # Define base quantities and calculate total quantities based on phantom count
    base_quantities = {
        "Main Part A": 936.5 * phantom_count,
        "Main Part B": 963.5 * phantom_count,
        "Microspheres Base": 9.5 * phantom_count,
        "Top Layer Part A": 13.4 * phantom_count,
        "Top Layer Part B": 1.6 * phantom_count,
        "Microspheres (Top Layer)": 0.075 * phantom_count,
        "Polane (Top Layer)": 7.5 * phantom_count,
        "Main Part A Contrast": 2.5 * phantom_count,
        "Main Part B Contrast": 2.5 * phantom_count,
        "Microspheres Target A": 0 * phantom_count,
        "Microspheres Target B": 0.008 * phantom_count,
        "Microspheres Target C": 0.075 * phantom_count,
        "Microspheres Target D": 0.65 * phantom_count
    }

    # Additional batch and ticket information from the right side of the screen
    batch_ticket_info = {
        "Serial Number": serial_number,
        "Main Part A Batch": data.get("mainPartABatch", ""),
        "Main Part A Ticket": data.get("mainPartATicket", ""),
        "Main Part B Batch": data.get("mainPartBBatch", ""),
        "Main Part B Ticket": data.get("mainPartBTicket", ""),
        "Top Layer Part A Batch": data.get("topLayerABatch", ""),
        "Top Layer Part A Ticket": data.get("topLayerATicket", ""),
        "Top Layer Part B Batch": data.get("topLayerBBatch", ""),
        "Top Layer Part B Ticket": data.get("topLayerBTicket", ""),
        "Microspheres Lot Number": data.get("microspheresLot", ""),
        "Microspheres Lot Reference": data.get("microspheresLotRef", "")
    }

    # Create a workbook and add both quantity and batch/ticket information
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Phantom Data"

    # Define header rows and add base quantities
    headers = ["Part", "Total Quantity (g/mL)"]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add base quantity data
    for part, quantity in base_quantities.items():
        worksheet.append([part, quantity])

    # Add an empty row and batch/ticket section
    worksheet.append([""])  # Empty row for spacing
    worksheet.append(["Batch and Ticket Information"])
    for key, value in batch_ticket_info.items():
        worksheet.append([key, value])

    # Apply styling to the Batch and Ticket Information header
    header_row = worksheet.max_row - len(batch_ticket_info) - 1
    worksheet[f"A{header_row}"].font = Font(bold=True, color="FFFFFF")
    worksheet[f"A{header_row}"].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Format cells: Set column width, border, and alignment
    for col in ["A", "B"]:
        worksheet.column_dimensions[col].width = 25

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save workbook to a BytesIO stream
    file_stream = BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    # Send the file as a response with a download prompt
    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"SerialNumber_{serial_number}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    app.run(debug=True)
